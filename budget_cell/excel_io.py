"""
FlatRow <-> Excel projection.

The projection (wide/long) is separated from extraction and normalization.
"""

from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path
from typing import Literal, Mapping

from budget_cell.types import FlatRow


OutputLayout = Literal["wide", "long"]
InputLayout = Literal["wide", "long", "flat"]


BASE_HEADERS: tuple[str, ...] = (
    "款", "項",
    "目", "本年度予算額", "前年度予算額", "比較",
    "国県支出金", "地方債", "その他", "一般財源",
    "節番号", "節名", "節金額",
    "小区分", "小区分金額",
    "事業コード",
)


_LEVEL_HEADER_RE = re.compile(r"^説明L(\d+)$")
_EMPTY_VALUES = (None, "")


def _to_text(value: object) -> str:
    return "" if value in _EMPTY_VALUES else str(value)


def _to_int(value: object) -> int | None:
    raw = (
        ""
        if value in _EMPTY_VALUES
        else (
            str(int(value))
            if isinstance(value, float)
            else str(value)
        )
    )
    normalized = raw.replace(",", "").replace("，", "").strip()
    return None if normalized == "" else int(normalized)


def _max_setsumei_level(rows: tuple[FlatRow, ...]) -> int:
    levels = tuple(
        row.setsumei_level
        for row in rows
        if row.setsumei_level is not None
    )
    return max(levels) if levels else 1


def _headers_wide(max_level: int) -> tuple[str, ...]:
    return (
        *BASE_HEADERS,
        *(f"説明L{i}" for i in range(1, max_level + 1)),
        "説明金額",
    )


def _headers_long() -> tuple[str, ...]:
    return (
        *BASE_HEADERS,
        "説明レベル",
        "説明",
        "説明金額",
    )


def _row_wide_tuple(row: FlatRow, max_level: int) -> tuple:
    level_cells = [""] * max_level
    if row.setsumei_name and row.setsumei_level is not None:
        idx = row.setsumei_level - 1
        if 0 <= idx < max_level:
            level_cells[idx] = row.setsumei_name

    return (
        row.kan_name,
        row.kou_name,
        row.moku_name,
        row.honendo if row.honendo is not None else "",
        row.zenendo if row.zenendo is not None else "",
        row.hikaku if row.hikaku is not None else "",
        row.kokuken if row.kokuken is not None else "",
        row.chihousei if row.chihousei is not None else "",
        row.sonota if row.sonota is not None else "",
        row.ippan if row.ippan is not None else "",
        row.setsu_number if row.setsu_number is not None else "",
        row.setsu_name,
        row.setsu_amount if row.setsu_amount is not None else "",
        row.sub_item_name,
        row.sub_item_amount if row.sub_item_amount is not None else "",
        row.setsumei_code,
        *level_cells,
        row.setsumei_amount if row.setsumei_amount is not None else "",
    )


def _row_long_tuple(row: FlatRow) -> tuple:
    return (
        row.kan_name,
        row.kou_name,
        row.moku_name,
        row.honendo if row.honendo is not None else "",
        row.zenendo if row.zenendo is not None else "",
        row.hikaku if row.hikaku is not None else "",
        row.kokuken if row.kokuken is not None else "",
        row.chihousei if row.chihousei is not None else "",
        row.sonota if row.sonota is not None else "",
        row.ippan if row.ippan is not None else "",
        row.setsu_number if row.setsu_number is not None else "",
        row.setsu_name,
        row.setsu_amount if row.setsu_amount is not None else "",
        row.sub_item_name,
        row.sub_item_amount if row.sub_item_amount is not None else "",
        row.setsumei_code,
        row.setsumei_level if row.setsumei_level is not None else "",
        row.setsumei_name,
        row.setsumei_amount if row.setsumei_amount is not None else "",
    )


def write_rows_to_excel_bytes(
    rows: tuple[FlatRow, ...],
    layout: OutputLayout = "wide",
) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Budget"

    max_level = _max_setsumei_level(rows) if layout == "wide" else 0
    headers = _headers_wide(max_level) if layout == "wide" else _headers_long()

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    for ci, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=ci, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 2):
        values = _row_wide_tuple(row, max_level) if layout == "wide" else _row_long_tuple(row)
        for ci, value in enumerate(values, 1):
            worksheet.cell(row=ri, column=ci, value=value)

    worksheet.auto_filter.ref = worksheet.dimensions
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def write_rows_to_excel_path(
    rows: tuple[FlatRow, ...],
    dst_path: str,
    layout: OutputLayout = "wide",
) -> None:
    Path(dst_path).write_bytes(write_rows_to_excel_bytes(rows, layout=layout))


def _header_map(headers: tuple[str, ...]) -> dict[str, int]:
    return {header: idx for idx, header in enumerate(headers) if header}


def _level_columns(header_to_index: Mapping[str, int]) -> tuple[tuple[int, int], ...]:
    parsed = tuple(
        (int(match.group(1)), idx)
        for header, idx in header_to_index.items()
        for match in (_LEVEL_HEADER_RE.match(header),)
        if match is not None
    )
    return tuple(sorted(parsed, key=lambda item: item[0]))


def _field_value(row: tuple[object, ...], header_to_index: Mapping[str, int], name: str) -> object:
    index = header_to_index.get(name)
    return row[index] if index is not None and index < len(row) else None


def _row_is_blank(row: tuple[object, ...]) -> bool:
    return all(value in _EMPTY_VALUES for value in row)


def _wide_setsumei(
    row: tuple[object, ...],
    level_cols: tuple[tuple[int, int], ...],
) -> tuple[int | None, str]:
    entries = tuple(
        (level, _to_text(row[idx]))
        for level, idx in level_cols
        if idx < len(row) and _to_text(row[idx]) != ""
    )
    return entries[0] if entries else (None, "")


def _long_setsumei(
    row: tuple[object, ...],
    header_to_index: Mapping[str, int],
) -> tuple[int | None, str]:
    return (
        _to_int(_field_value(row, header_to_index, "説明レベル")),
        _to_text(_field_value(row, header_to_index, "説明")),
    )


def _flat_setsumei(
    row: tuple[object, ...],
    header_to_index: Mapping[str, int],
) -> tuple[int | None, str]:
    return (None, _to_text(_field_value(row, header_to_index, "説明")))


def _layout_for_headers(
    header_to_index: Mapping[str, int],
    level_cols: tuple[tuple[int, int], ...],
) -> InputLayout:
    return (
        "long"
        if "説明レベル" in header_to_index and "説明" in header_to_index
        else "wide"
        if level_cols
        else "flat"
    )


def _row_from_excel(
    row: tuple[object, ...],
    header_to_index: Mapping[str, int],
    layout: InputLayout,
    level_cols: tuple[tuple[int, int], ...],
) -> FlatRow:
    setsumei_level, setsumei_name = (
        _long_setsumei(row, header_to_index)
        if layout == "long"
        else _wide_setsumei(row, level_cols)
        if layout == "wide"
        else _flat_setsumei(row, header_to_index)
    )
    return FlatRow(
        kan_name=_to_text(_field_value(row, header_to_index, "款")),
        kou_name=_to_text(_field_value(row, header_to_index, "項")),
        moku_name=_to_text(_field_value(row, header_to_index, "目")),
        honendo=_to_int(_field_value(row, header_to_index, "本年度予算額")),
        zenendo=_to_int(_field_value(row, header_to_index, "前年度予算額")),
        hikaku=_to_int(_field_value(row, header_to_index, "比較")),
        kokuken=_to_int(_field_value(row, header_to_index, "国県支出金")),
        chihousei=_to_int(_field_value(row, header_to_index, "地方債")),
        sonota=_to_int(_field_value(row, header_to_index, "その他")),
        ippan=_to_int(_field_value(row, header_to_index, "一般財源")),
        setsu_number=_to_int(_field_value(row, header_to_index, "節番号")),
        setsu_name=_to_text(_field_value(row, header_to_index, "節名")),
        setsu_amount=_to_int(_field_value(row, header_to_index, "節金額")),
        sub_item_name=_to_text(_field_value(row, header_to_index, "小区分")),
        sub_item_amount=_to_int(_field_value(row, header_to_index, "小区分金額")),
        setsumei_code=_to_text(_field_value(row, header_to_index, "事業コード")),
        setsumei_level=setsumei_level,
        setsumei_name=setsumei_name,
        setsumei_amount=_to_int(_field_value(row, header_to_index, "説明金額")),
    )


def read_rows_from_excel_bytes(data: bytes) -> tuple[FlatRow, ...]:
    import openpyxl

    workbook = openpyxl.load_workbook(BytesIO(data), data_only=True)
    worksheet = workbook.active
    header_row = next(
        worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
        (),
    )
    headers = tuple(_to_text(value).strip() for value in header_row)
    header_to_index = _header_map(headers)
    level_cols = _level_columns(header_to_index)
    layout = _layout_for_headers(header_to_index, level_cols)
    rows = tuple(worksheet.iter_rows(min_row=2, values_only=True))
    return tuple(
        _row_from_excel(row, header_to_index, layout, level_cols)
        for row in rows
        if not _row_is_blank(row)
    )


def read_rows_from_excel_path(src_path: str) -> tuple[FlatRow, ...]:
    return read_rows_from_excel_bytes(Path(src_path).read_bytes())
