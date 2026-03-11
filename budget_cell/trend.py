"""
Year-over-year trend representation from Excel-derived FlatRow streams.

This module is intentionally decoupled from PDF extraction.
Input contract: FlatRow sequence only.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Callable, Mapping, Sequence

from budget_cell.types import FlatRow


@dataclass(frozen=True)
class TrendKey:
    kan_name: str
    kou_name: str
    moku_name: str
    node_kind: str
    path_levels: tuple[str, ...]


@dataclass(frozen=True)
class TrendNode:
    year: str
    key: TrendKey
    setsu_number: int | None
    setsu_name: str
    sub_item_name: str
    setsumei_code: str
    setsumei_level: int
    item_name: str
    amount: int


@dataclass(frozen=True)
class TrendRow:
    key: TrendKey
    year_amounts: tuple[int, ...]
    diff: int
    ratio: float | None
    status: str


MatchIdFn = Callable[[TrendKey], str]

_NODE_KIND_ORDER = {
    "節": 0,
    "小区分": 1,
    "説明": 2,
}


def _year_sort_key(value: str) -> tuple[int, str]:
    match = re.search(r"(\d+)", value)
    return ((int(match.group(1)) if match else 10_000), value)


def _clean_text(value: str) -> str:
    return value.strip()


def _level_or_default(value: int | None) -> int:
    return value if value and value > 0 else 1


def _setsu_label(number: int | None, name: str) -> str:
    normalized_name = _clean_text(name)
    return (
        f"{number} {normalized_name}"
        if number is not None and normalized_name
        else str(number)
        if number is not None
        else normalized_name
    )


def _advance_path(
    previous: tuple[str, ...],
    level: int,
    name: str,
) -> tuple[str, ...]:
    size = max(len(previous), level)
    padded = tuple(
        previous[idx] if idx < len(previous) else ""
        for idx in range(size)
    )
    return tuple(
        name if idx == level - 1 else padded[idx] if idx < level - 1 else ""
        for idx in range(size)
    )


def _path_from_levels(levels: tuple[str, ...]) -> tuple[str, ...]:
    return tuple(item for item in levels if item)


def rows_to_trend_nodes(
    year: str,
    rows: Sequence[FlatRow],
) -> tuple[TrendNode, ...]:
    path_state: dict[tuple[str, str, str, int | None, str, str], tuple[str, ...]] = {}
    setsu_nodes: dict[tuple[str, str, str, int | None, str], TrendNode] = {}
    sub_item_nodes: dict[tuple[str, str, str, int | None, str, str], TrendNode] = {}
    setsumei_nodes: list[TrendNode] = []
    for row in rows:
        kan_name = _clean_text(row.kan_name)
        kou_name = _clean_text(row.kou_name)
        moku_name = _clean_text(row.moku_name)
        setsu_name = _clean_text(row.setsu_name)
        sub_item_name = _clean_text(row.sub_item_name)
        name = _clean_text(row.setsumei_name)
        level = _level_or_default(row.setsumei_level)
        setsu_path = tuple((label,)) if (label := _setsu_label(row.setsu_number, setsu_name)) else ()
        context_key = (
            kan_name,
            kou_name,
            moku_name,
            row.setsu_number,
            setsu_name,
            _clean_text(row.setsumei_code),
        )
        _ = (
            setsu_nodes.setdefault(
                (kan_name, kou_name, moku_name, row.setsu_number, setsu_name),
                TrendNode(
                    year=year,
                    key=TrendKey(
                        kan_name=kan_name,
                        kou_name=kou_name,
                        moku_name=moku_name,
                        node_kind="節",
                        path_levels=setsu_path,
                    ),
                    setsu_number=row.setsu_number,
                    setsu_name=setsu_name,
                    sub_item_name="",
                    setsumei_code="",
                    setsumei_level=1,
                    item_name=setsu_path[0] if setsu_path else "",
                    amount=row.setsu_amount,
                ),
            )
            if setsu_path and row.setsu_amount is not None
            else None
        )
        _ = (
            sub_item_nodes.setdefault(
                (kan_name, kou_name, moku_name, row.setsu_number, setsu_name, sub_item_name),
                TrendNode(
                    year=year,
                    key=TrendKey(
                        kan_name=kan_name,
                        kou_name=kou_name,
                        moku_name=moku_name,
                        node_kind="小区分",
                        path_levels=(*setsu_path, sub_item_name),
                    ),
                    setsu_number=row.setsu_number,
                    setsu_name=setsu_name,
                    sub_item_name=sub_item_name,
                    setsumei_code="",
                    setsumei_level=len((*setsu_path, sub_item_name)),
                    item_name=sub_item_name,
                    amount=row.sub_item_amount,
                ),
            )
            if sub_item_name and row.sub_item_amount is not None
            else None
        )
        next_levels = (
            _advance_path(path_state.get(context_key, ()), level, name)
            if name
            else path_state.get(context_key, ())
        )
        path_state[context_key] = next_levels
        amount = row.setsumei_amount
        setsumei_nodes.append(
            TrendNode(
                year=year,
                key=TrendKey(
                    kan_name=context_key[0],
                    kou_name=context_key[1],
                    moku_name=context_key[2],
                    node_kind="説明",
                    path_levels=_path_from_levels(next_levels),
                ),
                setsu_number=row.setsu_number,
                setsu_name=context_key[4],
                sub_item_name="",
                setsumei_code=context_key[5],
                setsumei_level=level,
                item_name=name,
                amount=amount if amount is not None else 0,
            )
        ) if name and amount is not None else None
    return (
        *tuple(setsu_nodes.values()),
        *tuple(sub_item_nodes.values()),
        *tuple(setsumei_nodes),
    )


def _status(base: int, latest: int) -> str:
    return (
        "新規"
        if base == 0 and latest != 0
        else "廃止"
        if base != 0 and latest == 0
        else "増額"
        if latest > base
        else "減額"
        if latest < base
        else "横ばい"
    )


def _ratio(base: int, diff: int) -> float | None:
    return None if base == 0 else diff / base


def trend_key_match_id_strict(key: TrendKey) -> str:
    return "|".join((key.kan_name, key.kou_name, key.moku_name, key.node_kind, *key.path_levels))


def _trend_key_sort_fields(key: TrendKey) -> tuple[str, str, str, int, tuple[str, ...]]:
    return (
        key.kan_name,
        key.kou_name,
        key.moku_name,
        _NODE_KIND_ORDER.get(key.node_kind, 99),
        key.path_levels,
    )


def _representative_key(variants: Mapping[TrendKey, int]) -> TrendKey:
    ranked = tuple(
        sorted(
            variants.items(),
            key=lambda item: (-item[1], _trend_key_sort_fields(item[0])),
        )
    )
    return ranked[0][0]


def aggregate_trends(
    nodes: Sequence[TrendNode],
    match_id_fn: MatchIdFn = trend_key_match_id_strict,
) -> tuple[tuple[str, ...], tuple[TrendRow, ...]]:
    years = tuple(sorted({node.year for node in nodes}, key=_year_sort_key))
    key_year_sum: dict[tuple[str, str], int] = {}
    key_variants: dict[str, dict[TrendKey, int]] = {}
    for node in nodes:
        match_id = match_id_fn(node.key)
        index = (match_id, node.year)
        key_year_sum[index] = key_year_sum.get(index, 0) + node.amount
        variants = key_variants.setdefault(match_id, {})
        variants[node.key] = variants.get(node.key, 0) + 1

    representative_keys = {
        match_id: _representative_key(variants)
        for match_id, variants in key_variants.items()
    }
    match_ids = tuple(
        sorted(
            representative_keys.keys(),
            key=lambda match_id: _trend_key_sort_fields(representative_keys[match_id]),
        )
    )
    keys = tuple(
        representative_keys[match_id]
        for match_id in match_ids
    )
    rows = tuple(
        (
            lambda amounts: TrendRow(
                key=key,
                year_amounts=amounts,
                diff=amounts[-1] - amounts[0],
                ratio=_ratio(amounts[0], amounts[-1] - amounts[0]),
                status=_status(amounts[0], amounts[-1]),
            )
        )(tuple(key_year_sum.get((match_id, year), 0) for year in years))
        for key, match_id in zip(keys, match_ids)
    )
    ranked = tuple(
        sorted(
            rows,
            key=lambda row: (abs(row.diff), row.key.kou_name, row.key.moku_name, row.key.path_levels),
            reverse=True,
        )
    )
    return years, ranked


def _max_path_depth(rows: Sequence[TrendRow]) -> int:
    return max((len(row.key.path_levels) for row in rows), default=1)


def _setsumei_headers(depth: int, years: Sequence[str]) -> tuple[str, ...]:
    return ("款", "項", "目", *(f"説明L{i}" for i in range(1, depth + 1)), *years, "増減額", "増減率", "状態", "増減額順位")


def _setsumei_sheet_row(
    row: TrendRow,
    depth: int,
    rank: int | None = None,
) -> tuple:
    levels = tuple(
        row.key.path_levels[idx] if idx < len(row.key.path_levels) else ""
        for idx in range(depth)
    )
    return (
        row.key.kan_name,
        row.key.kou_name,
        row.key.moku_name,
        *levels,
        *row.year_amounts,
        row.diff,
        row.ratio if row.ratio is not None else "",
        row.status,
        rank if rank is not None else "",
    )


def _setsu_headers(years: Sequence[str]) -> tuple[str, ...]:
    return ("款", "項", "目", "節", "小区分", *years, "増減額", "増減率", "状態", "増減額順位")


def _setsu_sheet_row(
    row: TrendRow,
    rank: int | None = None,
) -> tuple:
    setsu_label = row.key.path_levels[0] if row.key.path_levels else ""
    sub_item_name = row.key.path_levels[1] if len(row.key.path_levels) > 1 else ""
    return (
        row.key.kan_name,
        row.key.kou_name,
        row.key.moku_name,
        setsu_label,
        sub_item_name,
        *row.year_amounts,
        row.diff,
        row.ratio if row.ratio is not None else "",
        row.status,
        rank if rank is not None else "",
    )


def _combined_headers(depth: int, years: Sequence[str]) -> tuple[str, ...]:
    return (
        "款",
        "項",
        "目",
        "種別",
        "節",
        "小区分",
        *(f"説明L{i}" for i in range(1, depth + 1)),
        *years,
        "増減額",
        "増減率",
        "状態",
        "増減額順位",
    )


def _combined_sheet_row(
    row: TrendRow,
    depth: int,
    rank: int | None = None,
) -> tuple:
    setsumei_levels = tuple(
        row.key.path_levels[idx]
        if row.key.node_kind == "説明" and idx < len(row.key.path_levels)
        else ""
        for idx in range(depth)
    )
    setsu_label = (
        row.key.path_levels[0]
        if row.key.node_kind in ("節", "小区分") and row.key.path_levels
        else ""
    )
    sub_item_name = (
        row.key.path_levels[1]
        if row.key.node_kind == "小区分" and len(row.key.path_levels) > 1
        else ""
    )
    return (
        row.key.kan_name,
        row.key.kou_name,
        row.key.moku_name,
        row.key.node_kind,
        setsu_label,
        sub_item_name,
        *setsumei_levels,
        *row.year_amounts,
        row.diff,
        row.ratio if row.ratio is not None else "",
        row.status,
        rank if rank is not None else "",
    )


def _selected_rows(
    rows: Sequence[TrendRow],
    node_kinds: frozenset[str],
) -> tuple[TrendRow, ...]:
    return tuple(row for row in rows if row.key.node_kind in node_kinds)


def _write_ranked_sheet(
    sheet,
    rows: Sequence[TrendRow],
    row_fn,
) -> None:
    for ri, row in enumerate(rows, 2):
        values = row_fn(row, rank=ri - 1)
        for ci, value in enumerate(values, 1):
            sheet.cell(row=ri, column=ci, value=value)


def write_trend_excel(
    dst_path: str,
    nodes: Sequence[TrendNode],
    top_n: int = 200,
    match_id_fn: MatchIdFn = trend_key_match_id_strict,
) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    years, rows = aggregate_trends(nodes, match_id_fn=match_id_fn)
    setsumei_rows = _selected_rows(rows, frozenset({"説明"}))
    setsu_rows = _selected_rows(rows, frozenset({"節", "小区分"}))
    setsumei_depth = _max_path_depth(setsumei_rows)
    setsumei_headers = _setsumei_headers(setsumei_depth, years)
    setsu_headers = _setsu_headers(years)
    combined_headers = _combined_headers(setsumei_depth, years)

    workbook = openpyxl.Workbook()
    sheet_setsumei_changes = workbook.active
    sheet_setsumei_changes.title = "setsumei_changes"
    sheet_setsumei_short = workbook.create_sheet("setsumei_short")
    sheet_setsumei_up = workbook.create_sheet("setsumei_top_up")
    sheet_setsumei_down = workbook.create_sheet("setsumei_top_down")
    sheet_setsu_changes = workbook.create_sheet("setsu_changes")
    sheet_setsu_short = workbook.create_sheet("setsu_short")
    sheet_setsu_up = workbook.create_sheet("setsu_top_up")
    sheet_setsu_down = workbook.create_sheet("setsu_top_down")
    sheet_combined_changes = workbook.create_sheet("combined_changes")
    sheet_combined_short = workbook.create_sheet("combined_short")
    sheet_combined_up = workbook.create_sheet("combined_top_up")
    sheet_combined_down = workbook.create_sheet("combined_top_down")
    sheet_raw_setsumei = workbook.create_sheet("raw_setsumei")
    sheet_raw_setsu = workbook.create_sheet("raw_setsu")

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

    def write_header(sheet, labels: Sequence[str]) -> None:
        for ci, label in enumerate(labels, 1):
            cell = sheet.cell(row=1, column=ci, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    write_header(sheet_setsumei_changes, setsumei_headers)
    _write_ranked_sheet(
        sheet_setsumei_changes,
        setsumei_rows,
        lambda row, rank: _setsumei_sheet_row(row, setsumei_depth, rank=rank),
    )
    write_header(sheet_setsumei_short, setsumei_headers)
    _write_ranked_sheet(
        sheet_setsumei_short,
        setsumei_rows[:top_n],
        lambda row, rank: _setsumei_sheet_row(row, setsumei_depth, rank=rank),
    )
    write_header(sheet_setsumei_up, setsumei_headers)
    _write_ranked_sheet(
        sheet_setsumei_up,
        tuple(row for row in setsumei_rows if row.diff > 0)[:top_n],
        lambda row, rank: _setsumei_sheet_row(row, setsumei_depth, rank=rank),
    )
    write_header(sheet_setsumei_down, setsumei_headers)
    _write_ranked_sheet(
        sheet_setsumei_down,
        tuple(row for row in setsumei_rows if row.diff < 0)[:top_n],
        lambda row, rank: _setsumei_sheet_row(row, setsumei_depth, rank=rank),
    )

    write_header(sheet_setsu_changes, setsu_headers)
    _write_ranked_sheet(sheet_setsu_changes, setsu_rows, _setsu_sheet_row)
    write_header(sheet_setsu_short, setsu_headers)
    _write_ranked_sheet(sheet_setsu_short, setsu_rows[:top_n], _setsu_sheet_row)
    write_header(sheet_setsu_up, setsu_headers)
    _write_ranked_sheet(
        sheet_setsu_up,
        tuple(row for row in setsu_rows if row.diff > 0)[:top_n],
        _setsu_sheet_row,
    )
    write_header(sheet_setsu_down, setsu_headers)
    _write_ranked_sheet(
        sheet_setsu_down,
        tuple(row for row in setsu_rows if row.diff < 0)[:top_n],
        _setsu_sheet_row,
    )

    write_header(sheet_combined_changes, combined_headers)
    _write_ranked_sheet(
        sheet_combined_changes,
        rows,
        lambda row, rank: _combined_sheet_row(row, setsumei_depth, rank=rank),
    )
    write_header(sheet_combined_short, combined_headers)
    _write_ranked_sheet(
        sheet_combined_short,
        rows[:top_n],
        lambda row, rank: _combined_sheet_row(row, setsumei_depth, rank=rank),
    )
    write_header(sheet_combined_up, combined_headers)
    _write_ranked_sheet(
        sheet_combined_up,
        tuple(row for row in rows if row.diff > 0)[:top_n],
        lambda row, rank: _combined_sheet_row(row, setsumei_depth, rank=rank),
    )
    write_header(sheet_combined_down, combined_headers)
    _write_ranked_sheet(
        sheet_combined_down,
        tuple(row for row in rows if row.diff < 0)[:top_n],
        lambda row, rank: _combined_sheet_row(row, setsumei_depth, rank=rank),
    )

    raw_setsumei_headers = (
        "年度", "款", "項", "目", "種別", "階層レベル", "項目", "項目パス", "金額",
        "節番号", "節名", "小区分", "事業コード",
    )
    write_header(sheet_raw_setsumei, raw_setsumei_headers)
    setsumei_nodes = tuple(node for node in nodes if node.key.node_kind == "説明")
    for ri, node in enumerate(setsumei_nodes, 2):
        values = (
            node.year,
            node.key.kan_name,
            node.key.kou_name,
            node.key.moku_name,
            node.key.node_kind,
            node.setsumei_level,
            node.item_name,
            " > ".join(node.key.path_levels),
            node.amount,
            node.setsu_number if node.setsu_number is not None else "",
            node.setsu_name,
            node.sub_item_name,
            node.setsumei_code,
        )
        for ci, value in enumerate(values, 1):
            sheet_raw_setsumei.cell(row=ri, column=ci, value=value)

    raw_setsu_headers = (
        "年度", "款", "項", "目", "種別", "節", "小区分", "金額",
        "節番号", "節名",
    )
    write_header(sheet_raw_setsu, raw_setsu_headers)
    setsu_nodes = tuple(node for node in nodes if node.key.node_kind in ("節", "小区分"))
    for ri, node in enumerate(setsu_nodes, 2):
        values = (
            node.year,
            node.key.kan_name,
            node.key.kou_name,
            node.key.moku_name,
            node.key.node_kind,
            node.key.path_levels[0] if node.key.path_levels else "",
            node.key.path_levels[1] if len(node.key.path_levels) > 1 else "",
            node.amount,
            node.setsu_number if node.setsu_number is not None else "",
            node.setsu_name,
        )
        for ci, value in enumerate(values, 1):
            sheet_raw_setsu.cell(row=ri, column=ci, value=value)

    for sheet, headers in (
        (sheet_setsumei_changes, setsumei_headers),
        (sheet_setsumei_short, setsumei_headers),
        (sheet_setsumei_up, setsumei_headers),
        (sheet_setsumei_down, setsumei_headers),
        (sheet_setsu_changes, setsu_headers),
        (sheet_setsu_short, setsu_headers),
        (sheet_setsu_up, setsu_headers),
        (sheet_setsu_down, setsu_headers),
        (sheet_combined_changes, combined_headers),
        (sheet_combined_short, combined_headers),
        (sheet_combined_up, combined_headers),
        (sheet_combined_down, combined_headers),
    ):
        ratio_col = len(headers) - 2
        for ri in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=ri, column=ratio_col)
            cell.number_format = "0.0%"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = "A2"

    for sheet in (sheet_raw_setsumei, sheet_raw_setsu):
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = "A2"
    Path(dst_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(dst_path)


def load_year_excel_nodes(
    year_to_path: Mapping[str, str],
    read_rows_fn,
) -> tuple[TrendNode, ...]:
    return tuple(
        node
        for year, path in year_to_path.items()
        for node in rows_to_trend_nodes(year, read_rows_fn(path))
    )
