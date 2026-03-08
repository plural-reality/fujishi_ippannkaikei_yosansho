"""
CLI: Multi-page PDF → Excel.

Pipeline:
  extract → filter(歳出) → grid → headers
    → group by (款, 項) section
    → per section: cells → merge → parse → flatten → ffill(moku/setsu)
    → label(kan/kou)  — structural, every row
    → concat → Excel

Usage:
  python -m budget_cell.cli.to_excel <input.pdf> <output.xlsx>
"""

from __future__ import annotations

import sys
from itertools import groupby

from budget_cell.extract import extract_all_geometries
from budget_cell.grid import build_grid, extract_expenditure_pages
from budget_cell.header import parse_page_header
from budget_cell.cells import assign_words_to_cells
from budget_cell.merge import merge_rows
from budget_cell.section import split_page_sections
from budget_cell.parse import parse_page_budget
from budget_cell.flatten import (
    HEADERS, FFILL_FIELDS,
    flatten_all_pages, label_section, ffill,
    row_to_tuple,
)
from budget_cell.types import Cell, Grid, PageGeometry, PageHeader


def _write_excel(rows: tuple, dst_path: str) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 2):
        t = row_to_tuple(row)
        for ci, val in enumerate(t, 1):
            ws.cell(row=ri, column=ci, value=val)

    ws.auto_filter.ref = ws.dimensions
    wb.save(dst_path)


def _process_section(
    header: PageHeader,
    cell_groups: tuple[tuple[Cell, ...], ...],
) -> tuple:
    """Process one (款, 項) section: parse → flatten → ffill → label."""
    budgets = tuple(
        parse_page_budget(cells)
        for cells in cell_groups
    )
    flat = flatten_all_pages(budgets)
    filled = ffill(flat, FFILL_FIELDS)
    return label_section(header, filled)


def process_pdf_to_excel(src_path: str, dst_path: str) -> None:
    print(f"Extracting geometries from {src_path}...")
    all_geoms = extract_all_geometries(src_path)
    print(f"  {len(all_geoms)} pages total")

    geoms = extract_expenditure_pages(all_geoms)
    print(f"  {len(geoms)} expenditure pages (after 歳出 title page)")

    print("Building grids + extracting headers...")
    grids = tuple(map(build_grid, geoms))
    headers = tuple(parse_page_header(g, gr) for g, gr in zip(geoms, grids))

    # Filter to pages with valid headers (= budget data pages)
    valid = tuple(
        (h, g, gr)
        for h, g, gr in zip(headers, geoms, grids)
        if h is not None
    )
    print(f"  {len(valid)} pages with 款/項 headers")

    # Build cells → merge → split mid-page transitions → collect (header, cells) segments
    segments: tuple[tuple[PageHeader, tuple[Cell, ...]], ...] = tuple(
        segment
        for h, g, gr in valid
        for segment in split_page_sections(h, merge_rows(assign_words_to_cells(g, gr)))
    )
    print(f"  {len(segments)} segments (after mid-page splits)")

    # Group consecutive segments by (款, 項)
    sections = tuple(
        (key_header, tuple(cells for _, cells in group_segs))
        for key_header, group_segs in (
            (key, tuple(grp))
            for key, grp in groupby(
                segments,
                key=lambda t: (t[0].kan_number, t[0].kan_name, t[0].kou_number, t[0].kou_name),
            )
        )
        for key_header in (PageHeader(*key_header),)
    )
    print(f"  {len(sections)} sections (款/項)")

    print("Processing sections...")
    all_rows = tuple(
        row
        for header, cell_groups in sections
        for row in _process_section(header, cell_groups)
    )
    print(f"  {len(all_rows)} rows")

    print(f"Writing Excel: {dst_path}")
    _write_excel(all_rows, dst_path)
    print("Done.")


def main() -> None:
    usage = "Usage: python -m budget_cell.cli.to_excel <input.pdf> <output.xlsx>"
    src = sys.argv[1] if len(sys.argv) > 1 else None
    dst = sys.argv[2] if len(sys.argv) > 2 else None

    _ = (
        process_pdf_to_excel(src, dst) if src and dst else
        (print(usage), sys.exit(1))
    )


if __name__ == "__main__":
    main()
